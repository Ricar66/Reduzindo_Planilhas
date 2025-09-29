from __future__ import annotations
import csv
import io
import unicodedata
from openpyxl import load_workbook
from thefuzz import process

# --- FUNÇÕES AUXILIARES ---

def _normalizar(s: str) -> str:
    """Limpa e padroniza uma string para comparação de similaridade."""
    s = (s or "").strip().lower()
    return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))

def _ler_planilha(bytes_data: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Lê um arquivo CSV ou XLSX e retorna os cabeçalhos e as linhas de dados."""
    ext = (filename or "").split(".")[-1].lower()
    header_original = []
    linhas_com_cabecalho_original = []

    if ext in ("csv", "txt"):
        data = bytes_data.decode("utf-8", errors="ignore")
        try:
            dialect = csv.Sniffer().sniff(data.splitlines()[0], delimiters=',;')
        except (csv.Error, IndexError):
            dialect = 'excel'
        
        reader = csv.reader(io.StringIO(data), dialect)
        linhas = list(reader)
        if not linhas: return [], []
        
        header_original = linhas[0]
        for row in linhas[1:]:
            if all(not cell for cell in row): continue
            linhas_com_cabecalho_original.append({header_original[i]: row[i] if i < len(row) else "" for i in range(len(header_original))})

    elif ext in ("xlsx", "xlsm"):
        bio = io.BytesIO(bytes_data)
        ws = load_workbook(bio, data_only=True).active
        
        header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header_original = [str(cell or "") for cell in next(header_iter, [])]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row): continue
            linhas_com_cabecalho_original.append({header_original[i]: (str(row[i]) if i < len(header_original) and row[i] is not None else "") for i in range(len(header_original))})
    
    else:
        raise ValueError(f"Formato de arquivo não suportado: .{ext}")

    return header_original, linhas_com_cabecalho_original

def _mapear_cabecalhos_inteligente(cabecalhos_planilha: list[str], mapa_sistema: dict[str, list[str]]) -> dict:
    """Cria um mapa de correspondência entre os cabeçalhos da planilha e os campos do sistema."""
    
    # Cria um dicionário de "busca reversa": {apelido: campo_do_sistema}
    opcoes_busca = {}
    for campo_sistema, apelidos in mapa_sistema.items():
        opcoes_busca[_normalizar(campo_sistema)] = campo_sistema
        for apelido in apelidos:
            opcoes_busca[_normalizar(apelido)] = campo_sistema
            
    mapa_automatico = {}
    for cabecalho in cabecalhos_planilha:
        # Encontra a melhor correspondência na nossa lista de opções de busca
        resultado = process.extractOne(_normalizar(cabecalho), opcoes_busca.keys(), score_cutoff=80)
        
        if resultado:
            melhor_match = resultado[0]
            campo_sistema_correspondente = opcoes_busca[melhor_match]
            mapa_automatico[cabecalho] = campo_sistema_correspondente
            
    return mapa_automatico

# --- FUNÇÕES PRINCIPAIS ---

def importar_generico(bytes_data: bytes, filename: str, mapa_sistema: dict[str, list[str]]) -> list[dict]:
    """
    Processo de importação principal, agora mais organizado.
    """
    campos_do_sistema = list(mapa_sistema.keys())
    cabecalhos_originais, linhas_brutas = _ler_planilha(bytes_data, filename)
    
    if not linhas_brutas:
        return []

    mapa_automatico = _mapear_cabecalhos_inteligente(cabecalhos_originais, mapa_sistema)
    
    saida_final = []
    for linha in linhas_brutas:
        # Inicia um novo registro com todos os campos do sistema vazios
        novo_registro = {campo: "" for campo in campos_do_sistema}
        
        # Preenche o novo registro usando o mapa automático
        for cabecalho_original, valor in linha.items():
            if cabecalho_original in mapa_automatico:
                campo_sistema = mapa_automatico[cabecalho_original]
                novo_registro[campo_sistema] = valor.strip()
                
        saida_final.append(novo_registro)
        
    return saida_final

def analisar_cabecalhos_planilha(bytes_data: bytes, filename: str, mapa_sistema: dict[str, list[str]]) -> list[dict]:
    """
    MODO DETETIVE APRIMORADO: Mostra qual campo do sistema foi encontrado e a pontuação de similaridade.
    """
    cabecalhos_originais, _ = _ler_planilha(bytes_data, filename)
    opcoes_busca = {}
    for campo_sistema, apelidos in mapa_sistema.items():
        opcoes_busca[_normalizar(campo_sistema)] = campo_sistema
        for apelido in apelidos:
            opcoes_busca[_normalizar(apelido)] = campo_sistema

    analise = []
    for cabecalho in cabecalhos_originais:
        resultado = process.extractOne(_normalizar(cabecalho), opcoes_busca.keys(), score_cutoff=80)
        if resultado:
            melhor_match, pontuacao = resultado
            campo_sistema = opcoes_busca[melhor_match]
            analise.append({
                "original": cabecalho,
                "match": campo_sistema,
                "pontuacao": pontuacao
            })
        else:
            analise.append({
                "original": cabecalho,
                "match": "NÃO ENCONTRADO",
                "pontuacao": 0
            })
            
    return analise