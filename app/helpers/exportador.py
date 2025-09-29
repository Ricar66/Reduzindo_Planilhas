from __future__ import annotations
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def gerar_xlsx(titulo: str, registros: list[dict], campos: list[tuple[str,str]]) -> BytesIO:
    """Gera um XLSX em memória a partir de registros."""
    wb = Workbook(); ws = wb.active; ws.title = (titulo or "Planilha")[:31]
    for j, (rotulo, _) in enumerate(campos, start=1):
        ws.cell(row=1, column=j, value=rotulo)
        ws.column_dimensions[get_column_letter(j)].width = max(12, len(rotulo) + 2)
    
    for i, r in enumerate(registros, start=2):
        for j, (_, key) in enumerate(campos, start=1):
            valor = r.get(key, "")
            # --- INÍCIO DA CORREÇÃO ---
            # Se o valor for uma lista, converte para um texto separado por vírgulas
            if isinstance(valor, list):
                valor = ", ".join(map(str, valor))
            # --- FIM DA CORREÇÃO ---
            ws.cell(row=i, column=j, value=valor)
            
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf
